import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ==================== Color palette ====================
GREEN_LIGHT = "E1F5EE"
GREEN_MID = "1D9E75"
GREEN_DARK = "0F6E56"
BLUE_LIGHT = "E6F1FB"
BLUE_MID = "378ADD"
BLUE_DARK = "185FA5"
AMBER_LIGHT = "FAEEDA"
AMBER_MID = "BA7517"
RED_LIGHT = "FCEBEB"
RED_MID = "A32D2D"
PURPLE_LIGHT = "EEEDFE"
PURPLE_MID = "534AB7"
GRAY_LIGHT = "F1EFE8"
GRAY_MID = "444441"
WHITE = "FFFFFF"
HEADER_BG = "2C2C2A"
HEADER_FG = "FFFFFF"

thin = Side(style="thin", color="D3D1C7")
border_all = Border(left=thin, right=thin, top=thin, bottom=thin)
border_bottom = Border(bottom=Side(style="medium", color="D3D1C7"))

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.fill = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
        cell.font = Font(name="微软雅黑", size=10, bold=True, color=HEADER_FG)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        cell.border = border_all

def style_cell(cell, font_size=10, bold=False, color=GRAY_MID, bg=WHITE, wrap=True):
    cell.font = Font(name="微软雅黑", size=font_size, bold=bold, color=color)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=wrap)
    cell.border = border_all

def auto_width(ws, min_w=10, max_w=45):
    for col in ws.columns:
        max_len = min_w
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                lines = str(cell.value).split("\n")
                for line in lines:
                    max_len = max(max_len, len(line) * 1.2)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_w)
    ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

# ========================================================
# SHEET 1: TCO 7层成本结构
# ========================================================
ws1 = wb.active
ws1.title = "TCO成本结构"
ws1.sheet_properties.tabColor = GREEN_MID

ws1.merge_cells("A1:E1")
title_cell = ws1["A1"]
title_cell.value = "表1：Agent TCO 七层成本结构"
title_cell.font = Font(name="微软雅黑", size=14, bold=True, color=GREEN_DARK)
title_cell.alignment = Alignment(horizontal="left", vertical="center")
ws1.row_dimensions[1].height = 32

ws1.merge_cells("A2:E2")
fml = ws1["A2"]
fml.value = "总公式：Agent TCO = L1+L2+L3+L4+L5+L6+L7  |  每层 Ln = 月固定成本 + (单位消耗量 × 单价)"
fml.font = Font(name="微软雅黑", size=10, italic=True, color="666666")
fml.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
ws1.row_dimensions[2].height = 28

# Headers
headers1 = ["层级", "成本项", "计算方式", "典型占比", "标签"]
for i, h in enumerate(headers1, 1):
    ws1.cell(row=3, column=i, value=h)
style_header(ws1, 3, 5)

data1 = [
    ["L1", "模型调用费", "月Token消耗 × Token单价\n(输入价×3≈输出价)", "15-25%", "⭐ 唯一被算的项"],
    ["L2", "算力基础设施", "GPU租赁/折旧 + 带宽 + 电费", "10-15%", "❌ 常被忽略"],
    ["L3", "数据准备与标注", "标注人力 + 清洗工具 + 知识库搭建", "15-25%", "⚠ 严重被低估"],
    ["L4", "Agent开发与调优", "Prompt工程 + 框架搭建 + 测试迭代", "10-20%", "✅ 部分被算"],
    ["L5", "系统集成", "ERP/CRM/API对接 + 网关建设", "10-15%", "❌ 常被忽略"],
    ["L6", "运维监控", "监控系统 + 告警 + 版本管理", "5-10%", "❌ 几乎未被算"],
    ["L7", "安全合规", "权限体系 + 审计日志 + 等保合规", "5-10%", "❌ 几乎未被算"],
]

for r, row_data in enumerate(data1, 4):
    for c, val in enumerate(row_data, 1):
        cell = ws1.cell(row=r, column=c, value=val)
        bg = WHITE
        if c == 5:
            if "唯一" in str(val):
                bg = GREEN_LIGHT
            elif "严重" in str(val):
                bg = RED_LIGHT
            elif "忽略" in str(val):
                bg = GRAY_LIGHT
        style_cell(cell, bg=bg)
    ws1.row_dimensions[r].height = 36

# Key insight row
ws1.merge_cells("A11:E11")
insight = ws1["A11"]
insight.value = "⚠ 关键洞察：只算Token单价来算ROI，误差可达4倍以上。L3-L7五层隐藏成本占总TCO的45-80%"
insight.font = Font(name="微软雅黑", size=10, bold=True, color=AMBER_MID)
insight.fill = PatternFill(start_color=AMBER_LIGHT, end_color=AMBER_LIGHT, fill_type="solid")
insight.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
insight.border = border_all
ws1.row_dimensions[11].height = 30

auto_width(ws1)
ws1.column_dimensions["C"].width = 30
ws1.column_dimensions["D"].width = 12

# ========================================================
# SHEET 2: Agent ROI公式
# ========================================================
ws2 = wb.create_sheet("Agent ROI公式")
ws2.sheet_properties.tabColor = "D85A30"

ws2.merge_cells("A1:D1")
t2 = ws2["A1"]
t2.value = "表2：Agent ROI 核心公式（展开式 + 变量定义 + 135倍杠杆推导）"
t2.font = Font(name="微软雅黑", size=14, bold=True, color="993C1D")
t2.alignment = Alignment(horizontal="left", vertical="center")
ws2.row_dimensions[1].height = 32

# Formula block
ws2.merge_cells("A2:D2")
f2 = ws2["A2"]
f2.value = "公式①：Agent ROI = (Agent产出价值 − Agent持有成本) / Agent持有成本 × 100%"
f2.font = Font(name="微软雅黑", size=11, bold=True, color=GRAY_MID)
f2.fill = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")
f2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
f2.border = border_all
ws2.row_dimensions[2].height = 28

ws2.merge_cells("A3:D3")
f3 = ws2["A3"]
f3.value = "公式②（展开）：Agent ROI = (∑ 月完成工单数 × 单工单价值 − TCO月均) / TCO月均 × 100%"
f3.font = Font(name="微软雅黑", size=11, bold=True, color=GRAY_MID)
f3.fill = PatternFill(start_color=GRAY_LIGHT, end_color=GRAY_LIGHT, fill_type="solid")
f3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
f3.border = border_all
ws2.row_dimensions[3].height = 28

# Variable table
headers2 = ["变量", "定义", "计算方式", "参考值"]
for i, h in enumerate(headers2, 1):
    ws2.cell(row=4, column=i, value=h)
style_header(ws2, 4, 4)

data2 = [
    ["产出价值", "Agent完成的业务产出", "完成工单数 × 单工单价值\n（单工单价值参照人类员工时薪折算）", "对标人类产值"],
    ["持有成本", "Agent月度TCO", "Σ(L1-L7) 月均成本\n（参考表1的7层成本结构累加）", "¥2,000-50,000/月"],
    ["产能复用率", "同时处理的任务线数", "Agent并发数 / 人类单线程数\n（核心杠杆：单条线负ROI→多线暴增）", "100-1000x"],
    ["产能利用率", "实际运行时间占比", "实际运行h / 总可用h × 100%\n人类含休息/开会/摸鱼时间", "人类≈60% / Agent≥95%"],
]

for r, row_data in enumerate(data2, 5):
    for c, val in enumerate(row_data, 1):
        cell = ws2.cell(row=r, column=c, value=val)
        bg = WHITE
        if r == 7:  # 产能复用率行高亮
            bg = AMBER_LIGHT
        style_cell(cell, bg=bg)
    ws2.row_dimensions[r].height = 42

# 135x lever
ws2.merge_cells("A9:D9")
lev = ws2["A9"]
lev.value = "━━━ 135x ROI杠杆 推导过程（产能复用率示例）━━━"
lev.font = Font(name="微软雅黑", size=11, bold=True, color=WHITE)
lev.fill = PatternFill(start_color=GREEN_MID, end_color=GREEN_MID, fill_type="solid")
lev.alignment = Alignment(horizontal="left", vertical="center")
lev.border = border_all
ws2.row_dimensions[9].height = 28

ws2.merge_cells("A10:D10")
d1 = ws2["A10"]
d1.value = "假设条件：Agent月持有成本=¥2,000 | 替代人类2天工作量（月薪¥15,000 × 2/22 ≈ ¥1,360/2天）"
d1.font = Font(name="微软雅黑", size=10, color=GRAY_MID)
d1.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
d1.border = border_all
ws2.row_dimensions[10].height = 24

ws2.merge_cells("A11:D11")
d2 = ws2["A11"]
d2.value = "单条线 ROI = (1,360 - 2,000) / 2,000 = −32%  ← 负值！单条线亏本"
d2.font = Font(name="微软雅黑", size=10, color="D85A30", bold=True)
d2.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
d2.border = border_all
ws2.row_dimensions[11].height = 24

ws2.merge_cells("A12:D12")
d3 = ws2["A12"]
d3.value = "100条线 ROI = (200 × 1,360 - 2,000) / 2,000 = 135x  ← 规模化复用让ROI暴增"
d3.font = Font(name="微软雅黑", size=12, color=GREEN_DARK, bold=True)
d3.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
d3.border = border_all
ws2.row_dimensions[12].height = 28

ws2.merge_cells("A13:D13")
d4 = ws2["A13"]
d4.value = "核心结论：Agent ROI的真正杠杆不在单任务提效，在规模化复用。当复用率≥50条线时，ROI从负值翻转为正值。"
d4.font = Font(name="微软雅黑", size=10, bold=True, color=BLUE_DARK)
d4.fill = PatternFill(start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type="solid")
d4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
d4.border = border_all
ws2.row_dimensions[13].height = 30

auto_width(ws2)
ws2.column_dimensions["A"].width = 14
ws2.column_dimensions["B"].width = 24
ws2.column_dimensions["C"].width = 35
ws2.column_dimensions["D"].width = 18

# ========================================================
# SHEET 3: 人机对照
# ========================================================
ws3 = wb.create_sheet("人机对照")
ws3.sheet_properties.tabColor = BLUE_MID

ws3.merge_cells("A1:D1")
t3 = ws3["A1"]
t3.value = "表3：人类员工 vs Agent员工 成本对照表"
t3.font = Font(name="微软雅黑", size=14, bold=True, color=BLUE_DARK)
t3.alignment = Alignment(horizontal="left", vertical="center")
ws3.row_dimensions[1].height = 32

headers3 = ["对比维度", "人类员工", "Agent员工", "差异倍数"]
for i, h in enumerate(headers3, 1):
    ws3.cell(row=2, column=i, value=h)
style_header(ws3, 2, 4)

data3 = [
    ["工作时长", "8h/天 · 5天/周", "24h/7天 · 无休", "4.2x"],
    ["产能复用", "1条任务线", "100-1000条并行", "100-1000x"],
    ["响应速度", "分钟-小时级", "毫秒-秒级", "60-3600x"],
    ["错误率", "1-5%（视疲劳程度）", "稳定 <0.1%", "10-50x"],
    ["学习成本", "入职培训1-3个月", "Prompt调试1-3天", "30x"],
    ["离职风险", "随时可能", "版本可控", "—"],
]

for r, row_data in enumerate(data3, 3):
    for c, val in enumerate(row_data, 1):
        cell = ws3.cell(row=r, column=c, value=val)
        bg = WHITE
        if c == 4 and val != "—":
            bg = GREEN_LIGHT
        style_cell(cell, bg=bg)
    ws3.row_dimensions[r].height = 26

ws3.merge_cells("A9:D9")
note3 = ws3["A9"]
note3.value = "※ 差异倍数取最保守估计。实际场景中Agent的价值放大效应在知识密集型岗位（如代码/设计/分析）更为显著"
note3.font = Font(name="微软雅黑", size=9, italic=True, color="888888")
note3.alignment = Alignment(horizontal="left", vertical="center")
note3.border = border_all

auto_width(ws3)
ws3.column_dimensions["A"].width = 14
ws3.column_dimensions["B"].width = 24
ws3.column_dimensions["C"].width = 24
ws3.column_dimensions["D"].width = 14

# ========================================================
# SHEET 4: 四阶段ROI
# ========================================================
ws4 = wb.create_sheet("四阶段ROI")
ws4.sheet_properties.tabColor = "639922"

ws4.merge_cells("A1:E1")
t4 = ws4["A1"]
t4.value = "表4：ROI四阶段计算对照表"
t4.font = Font(name="微软雅黑", size=14, bold=True, color="27500A")
t4.alignment = Alignment(horizontal="left", vertical="center")
ws4.row_dimensions[1].height = 32

headers4 = ["阶段", "时间窗口", "计算公式", "关键指标", "通过标准"]
for i, h in enumerate(headers4, 1):
    ws4.cell(row=2, column=i, value=h)
style_header(ws4, 2, 5)

data4 = [
    ["实验期", "0-3个月", "不算钱，算可行性", "Agent任务完成率\n人工审核通过率\n用户采纳率", "完成率>80%\n通过率>70%\n采纳率>60%"],
    ["试点期", "3-6个月", "AB组对比法\nη=(Tₐ−Tₐᵢ)/Tₐ\nTₐ=人工耗时 Tₐᵢ=Agent辅助耗时", "A组(人工+Agent) vs B组(纯人工)\n单位耗时·出错率·满意度", "效率提升>30%\n或 错误下降>50%"],
    ["规模化期", "6-12个月", "ROI = (收益−TCO)/TCO", "单Agent成本\n投产比", "优秀>300%\n良好100-300%\n及格30-100%"],
    ["深度嵌入期", "12个月+", "经营指标归因\n（非直接公式）", "客户留存率·人效比\n业务响应周期·创新转化率", "趋势持续向好"],
]

for r, row_data in enumerate(data4, 3):
    for c, val in enumerate(row_data, 1):
        cell = ws4.cell(row=r, column=c, value=val)
        bg = WHITE
        if r == 3:
            bg = GREEN_LIGHT
        elif r == 4:
            bg = AMBER_LIGHT
        elif r == 5:
            bg = RED_LIGHT
        elif r == 6:
            bg = PURPLE_LIGHT
        style_cell(cell, bg=bg)
    ws4.row_dimensions[r].height = 56

ws4.merge_cells("A7:E7")
note4 = ws4["A7"]
note4.value = "示例：审核1单从120分钟→30分钟，η=(120-30)/120=75%，效率提升75% > 30% → 试点期通过"
note4.font = Font(name="微软雅黑", size=10, italic=True, color=BLUE_DARK)
note4.fill = PatternFill(start_color=BLUE_LIGHT, end_color=BLUE_LIGHT, fill_type="solid")
note4.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
note4.border = border_all
ws4.row_dimensions[7].height = 28

auto_width(ws4)
ws4.column_dimensions["C"].width = 32
ws4.column_dimensions["D"].width = 28
ws4.column_dimensions["E"].width = 20

# ========================================================
# SHEET 5: 行业对标
# ========================================================
ws5 = wb.create_sheet("行业对标")
ws5.sheet_properties.tabColor = "993556"

ws5.merge_cells("A1:E1")
t5 = ws5["A1"]
t5.value = "表5：行业Agent ROI 对标基准表"
t5.font = Font(name="微软雅黑", size=14, bold=True, color="72243E")
t5.alignment = Alignment(horizontal="left", vertical="center")
ws5.row_dimensions[1].height = 32

headers5 = ["行业", "典型Agent场景", "ROI参考区间", "关键决定因素", "标杆案例"]
for i, h in enumerate(headers5, 1):
    ws5.cell(row=2, column=i, value=h)
style_header(ws5, 2, 5)

data5 = [
    ["电商/零售", "虚拟试衣·客服·推荐", "200-500%", "退货率降幅·客单价提升", "安踏"],
    ["餐饮", "门店巡检·客服", "200-400%", "人力节省·满意度", "海底捞"],
    ["汽车", "HMI设计·仿真", "300-500%", "研发周期压缩", "赛力斯"],
    ["制造业", "质检·排产·仿真", "150-400%", "不良品降幅·设备利用率", "—"],
    ["金融", "风控·投研·客服", "100-300%", "错误减少·处理速度", "—"],
    ["互联网", "代码生成·内容生产", "300-800%", "产能复用率（核心杠杆）", "字节"],
]

for r, row_data in enumerate(data5, 3):
    for c, val in enumerate(row_data, 1):
        cell = ws5.cell(row=r, column=c, value=val)
        bg = WHITE
        if c == 3:
            pct = str(val).replace("%", "").split("-")
            try:
                if int(pct[0]) >= 300:
                    bg = GREEN_LIGHT
                elif int(pct[0]) >= 200:
                    bg = AMBER_LIGHT
                else:
                    bg = RED_LIGHT
            except:
                pass
        style_cell(cell, bg=bg)
    ws5.row_dimensions[r].height = 24

ws5.merge_cells("A9:E9")
note5 = ws5["A9"]
note5.value = "※ ROI区间为规模化期（6-12月）参考值，实际因企业规模、数据基础、场景复杂度而异"
note5.font = Font(name="微软雅黑", size=9, italic=True, color="888888")
note5.alignment = Alignment(horizontal="left", vertical="center")
note5.border = border_all

auto_width(ws5)

# ========================================================
# SAVE
# ========================================================
output_path = r"C:\Users\jt\WorkBuddy\Claw\Agent_ROI计算表.xlsx"
wb.save(output_path)
print(f"Saved: {output_path}")
